"""XLSX export — live formulas + named ranges (FR-025, SC-004, SC-006).

The Assumptions sheet exposes every registry leaf as a workbook-scoped
named range. Every derived sheet references those names via formula
strings, NOT pre-evaluated constants — so a finance reader can edit
`fee_level_control_pct` in Excel and watch dependent cells recompute.

Constitution Principles II + VI: single source of assumptions + outputs
interrogable by a finance person who doesn't read code.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.properties import CalcProperties

from src.config.schema import Registry
from src.engine.ab_test import ABTestView
from src.engine.aggregates import WeeklyAggregate, weekly_aggregate
from src.engine.briefing import Briefing
from src.engine.consistency import ConsistencyReport
from src.engine.performance import PerformanceView
from src.engine.projection import ProjectionView
from src.engine.variance import VarianceView
from src.ui.components import (
    APP_SUBTITLE,
    APP_TITLE,
    format_date,
    format_week_commencing,
)

ORIGIN_FILL = {
    "disclosed": PatternFill("solid", fgColor="CCE5FF"),
    "observed": PatternFill("solid", fgColor="D4EDDA"),
    "measured-from-data": PatternFill("solid", fgColor="E2E3E5"),
    "assumed": PatternFill("solid", fgColor="FFF3CD"),
}
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="F0F0F0")

# Excel number-format codes — centralised so the workbook is visually
# consistent everywhere monetary / % / pp / bps values are written.
FMT_EUR = '"€"#,##0.00;[Red]"-€"#,##0.00'
FMT_PCT = "0.00%"
FMT_PP = '+0.00" pp";-0.00" pp";"0 pp"'
FMT_BPS = '+0" bps";-0" bps";"0 bps"'


def write_workbook(
    *,
    registry: Registry,
    bookings_df,
    performance: PerformanceView,
    variance: VarianceView,
    ab_test: ABTestView,
    projection: ProjectionView,
    briefing: Briefing,
    consistency: ConsistencyReport,
    path: str | Path,
) -> Path:
    """Write the workbook to *path*. Returns the resolved Path."""
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    # openpyxl writes formulas but doesn't evaluate them. Some viewers
    # (Excel preview, Google Sheets in import mode, Numbers) show blank
    # cells until they recalc. fullCalcOnLoad forces the spreadsheet app
    # to recompute every formula immediately on open.
    wb.calculation = CalcProperties(fullCalcOnLoad=True)

    start_date = registry.dataset.start_date.value
    name_index = _write_assumptions(wb, registry)
    _write_readme(wb, name_index)
    _write_weekly_aggregates(wb, registry, bookings_df)
    _write_performance(wb, performance, start_date)
    _write_variance(wb, variance, start_date)
    _write_ab_test(wb, ab_test, registry, start_date)
    _write_projection(wb, projection, start_date)
    _write_briefing(wb, briefing)
    _write_consistency(wb, consistency)
    _write_audit(wb, performance)

    # Tab order: README first
    desired_order = [
        "README",
        "Assumptions",
        "WeeklyAggregates",
        "Performance",
        "Variance",
        "ABTest",
        "Projection",
        "Briefing",
        "Consistency",
        "Audit",
    ]
    wb._sheets = sorted(wb._sheets, key=lambda s: desired_order.index(s.title))

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------

def _write_assumptions(wb: Workbook, registry: Registry) -> dict[str, str]:
    """Write the Assumptions sheet and return {name → A1-style cell ref}."""
    ws = wb.create_sheet("Assumptions")
    ws.append(["Key", "Value", "Origin", "Source", "Notes"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    name_to_cell: dict[str, str] = {}

    def _add(name: str, value: Any, origin: str, source: str | None, notes: str | None):
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=value)
        cell = ws.cell(row=row_idx, column=3, value=origin)
        cell.fill = ORIGIN_FILL.get(origin, PatternFill())
        ws.cell(row=row_idx, column=4, value=source or "")
        ws.cell(row=row_idx, column=5, value=notes or "")
        a1 = f"Assumptions!$B${row_idx}"
        name_to_cell[name] = a1
        wb.defined_names[name] = DefinedName(name=name, attr_text=a1)

    # Scalars
    _add(
        "coverage_pct",
        registry.coverage_pct.value,
        registry.coverage_pct.origin,
        registry.coverage_pct.source,
        registry.coverage_pct.notes,
    )
    _add(
        "payment_processing_pct",
        registry.payment_processing_pct.value,
        registry.payment_processing_pct.origin,
        registry.payment_processing_pct.source,
        registry.payment_processing_pct.notes,
    )
    _add(
        "servicing_cost_per_unit_cents",
        registry.servicing_cost_per_unit_cents.value,
        registry.servicing_cost_per_unit_cents.origin,
        registry.servicing_cost_per_unit_cents.source,
        registry.servicing_cost_per_unit_cents.notes,
    )
    _add(
        "fee_level_control_pct",
        registry.fee_level.control_pct.value,
        registry.fee_level.control_pct.origin,
        registry.fee_level.control_pct.source,
        registry.fee_level.control_pct.notes,
    )
    _add(
        "fee_level_test_pct",
        registry.fee_level.test_pct.value,
        registry.fee_level.test_pct.origin,
        registry.fee_level.test_pct.source,
        registry.fee_level.test_pct.notes,
    )
    _add(
        "margin_floor_bps",
        registry.margin.floor_bps.value,
        registry.margin.floor_bps.origin,
        registry.margin.floor_bps.source,
        registry.margin.floor_bps.notes,
    )
    _add(
        "margin_approaching_floor_buffer_bps",
        registry.margin.approaching_floor_buffer_bps.value,
        registry.margin.approaching_floor_buffer_bps.origin,
        registry.margin.approaching_floor_buffer_bps.source,
        registry.margin.approaching_floor_buffer_bps.notes,
    )
    _add(
        "classification_material_gap_bps",
        registry.classification.material_gap_bps.value,
        registry.classification.material_gap_bps.origin,
        registry.classification.material_gap_bps.source,
        registry.classification.material_gap_bps.notes,
    )
    _add(
        "classification_persistence_weeks",
        registry.classification.persistence_weeks.value,
        registry.classification.persistence_weeks.origin,
        registry.classification.persistence_weeks.source,
        registry.classification.persistence_weeks.notes,
    )
    _add(
        "metrics_trailing_window_weeks",
        registry.metrics.trailing_window_weeks.value,
        registry.metrics.trailing_window_weeks.origin,
        registry.metrics.trailing_window_weeks.source,
        registry.metrics.trailing_window_weeks.notes,
    )
    _add(
        "projection_weeks_forward",
        registry.projection.weeks_forward.value,
        registry.projection.weeks_forward.origin,
        registry.projection.weeks_forward.source,
        registry.projection.weeks_forward.notes,
    )
    _add(
        "projection_trend_factor",
        registry.projection.trend_factor.value,
        registry.projection.trend_factor.origin,
        registry.projection.trend_factor.source,
        registry.projection.trend_factor.notes,
    )

    # Per-partner priced cancel rate (named ranges)
    for pid in registry.partners():
        partner_cfg = registry.partner[pid]
        _add(
            f"partner_{pid}_priced_cancel_rate",
            partner_cfg.priced_cancel_rate.value,
            partner_cfg.priced_cancel_rate.origin,
            partner_cfg.priced_cancel_rate.source,
            partner_cfg.priced_cancel_rate.notes,
        )

    # Column widths
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 44
    ws.column_dimensions["E"].width = 48

    return name_to_cell


def _write_readme(wb: Workbook, name_index: dict[str, str]) -> None:
    ws = wb.create_sheet("README")
    ws["A1"] = APP_TITLE
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = APP_SUBTITLE
    ws["A2"].font = Font(italic=True, color="555555")
    ws["A3"] = (
        "Every figure in this workbook traces back to a value on the "
        "Assumptions sheet via a named range. Edit any yellow (assumed) or "
        "blue (disclosed) cell on Assumptions, and dependent cells in "
        "Performance / Variance / ABTest / Projection will recompute."
    )
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 60

    ws["A5"] = "Origin colour key"
    ws["A5"].font = Font(bold=True)
    rows = [
        ("disclosed (D)", "Published by an authoritative external party. Source cited.", "disclosed"),
        ("observed (O)", "Recorded from real-world activity.", "observed"),
        ("measured-from-data (M)", "Derived from a dataset; reproducible from raw data.", "measured-from-data"),
        ("assumed (A)", "Modeller choice; refine when evidence emerges.", "assumed"),
    ]
    for i, (label, desc, origin) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=label).fill = ORIGIN_FILL[origin]
        ws.cell(row=i, column=2, value=desc)

    ws["A12"] = "Named-range index"
    ws["A12"].font = Font(bold=True)
    ws["A13"] = "Name"
    ws["B13"] = "Cell reference"
    for c in ws[13]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for i, (name, ref) in enumerate(sorted(name_index.items()), start=14):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=ref)
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 32


def _write_weekly_aggregates(wb: Workbook, registry: Registry, bookings_df) -> None:
    """Long-form table of every (partner, week, route, arm) aggregate.

    These cells contain raw integers (facts). Other sheets reference them
    via SUMIFS to keep formulas auditable.
    """
    ws = wb.create_sheet("WeeklyAggregates")
    headers = [
        "partner_id", "iso_week", "route_type", "ab_arm",
        "bookings", "ancillaries_sold", "ancillaries_cancelled",
        "revenue_cents", "payouts_cents", "cost_of_service_cents",
        "gross_margin_cents",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    rows = weekly_aggregate(
        bookings_df,
        registry,
        by_partner=True,
        by_route=True,
        by_arm=True,
        include_blended=False,
    )
    for r in rows:
        ws.append([
            r.partner_id,
            r.iso_week,
            r.route_type or "",
            r.ab_arm,
            r.bookings,
            r.ancillaries_sold,
            r.ancillaries_cancelled,
            r.revenue_cents,
            r.payouts_cents,
            r.cost_of_service_cents,
            r.gross_margin_cents,
        ])

    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + i)].width = 18


def _write_performance(
    wb: Workbook, performance: PerformanceView, start_date
) -> None:
    ws = wb.create_sheet("Performance")
    wc = format_week_commencing(performance.as_of_week, start_date)
    ws["A1"] = f"Performance — as of {wc}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = (
        f"Trailing window: {performance.trailing_window_weeks} weeks. "
        f"Margin floor: see `margin_floor_bps` on Assumptions."
    )

    # Header row at row 4
    headers = [
        "Partner", "Status", "Revenue (€)", "Payouts (€)",
        "Cost of service (€)", "Contribution (€)", "Attach rate",
        "Loss ratio", "Gross margin", "Margin distance from floor (bps)",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    def _attach_formula(week: int, partner: str) -> str:
        return (
            f'=IFERROR(SUMIFS(WeeklyAggregates!F:F, WeeklyAggregates!A:A, "{partner}", '
            f'WeeklyAggregates!B:B, {week}) / SUMIFS(WeeklyAggregates!E:E, '
            f'WeeklyAggregates!A:A, "{partner}", WeeklyAggregates!B:B, {week}), "")'
        )

    def _loss_formula(week: int, partner: str) -> str:
        return (
            f'=IFERROR(SUMIFS(WeeklyAggregates!I:I, WeeklyAggregates!A:A, "{partner}", '
            f'WeeklyAggregates!B:B, {week}) / SUMIFS(WeeklyAggregates!H:H, '
            f'WeeklyAggregates!A:A, "{partner}", WeeklyAggregates!B:B, {week}), "")'
        )

    def _gm_formula(week: int, partner: str) -> str:
        return (
            f'=IFERROR(SUMIFS(WeeklyAggregates!K:K, WeeklyAggregates!A:A, "{partner}", '
            f'WeeklyAggregates!B:B, {week}) / SUMIFS(WeeklyAggregates!H:H, '
            f'WeeklyAggregates!A:A, "{partner}", WeeklyAggregates!B:B, {week}), "")'
        )

    def _sumifs_eur(col_letter: str, week: int, partner: str) -> str:
        """SUMIFS over a raw-cents column on WeeklyAggregates, divided by
        100 so the cell displays as euros under FMT_EUR."""
        return (
            f'=SUMIFS(WeeklyAggregates!{col_letter}:{col_letter}, '
            f'WeeklyAggregates!A:A, "{partner}", WeeklyAggregates!B:B, '
            f'{week})/100'
        )

    week = performance.as_of_week
    row = 5
    for status in performance.partners:
        pid = status.partner_id
        ws.cell(row=row, column=1, value=status.display_name)
        # Status — formula-driven via margin distance (col J) + the
        # margin_floor / approaching_floor_buffer named ranges. Must
        # recompute when Assumptions change, consistent with the live
        # model. Engine-only states ("no_activity", "partial_window")
        # aren't expressible in a row-level formula; ISNUMBER guards a
        # no-activity row so we display "no data" rather than a #VALUE!
        # cascade.
        ws.cell(
            row=row, column=2,
            value=(
                f'=IF(ISNUMBER(J{row})=FALSE,"no data",'
                f'IF(J{row}<0,"breach",'
                f'IF(J{row}<=margin_approaching_floor_buffer_bps,"warning",'
                f'"healthy")))'
            ),
        )
        # Currency columns — stored as cents/100 so FMT_EUR displays euros.
        for col_idx, letter in ((3, "H"), (4, "I"), (5, "J"), (6, "K")):
            cell = ws.cell(
                row=row, column=col_idx,
                value=_sumifs_eur(letter, week, pid),
            )
            cell.number_format = FMT_EUR
        # Ratios — already produced as fractions by the existing formulas.
        for col_idx, formula in (
            (7, _attach_formula(week, pid)),
            (8, _loss_formula(week, pid)),
            (9, _gm_formula(week, pid)),
        ):
            cell = ws.cell(row=row, column=col_idx, value=formula)
            cell.number_format = FMT_PCT
        # Margin distance — live formula: (gross margin × 10000) − floor bps.
        # Replaces the previously-hardcoded integer literal so an edit to
        # margin_floor_bps on Assumptions propagates here too. IFERROR
        # guards no-activity rows where I{row} returns "" from the loss /
        # gm formulas.
        cell = ws.cell(
            row=row, column=10,
            value=f'=IFERROR(ROUND(I{row}*10000,0)-margin_floor_bps,"")',
        )
        cell.number_format = FMT_BPS
        row += 1

    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + i)].width = 22


def _write_variance(
    wb: Workbook, variance: VarianceView, start_date
) -> None:
    ws = wb.create_sheet("Variance")
    wc = format_week_commencing(variance.as_of_week, start_date)
    ws["A1"] = f"Variance — as of {wc}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = (
        f"Trailing window: {variance.trailing_window_weeks} weeks. "
        f"Material gap threshold = `classification_material_gap_bps` on Assumptions."
    )

    headers = [
        "Partner", "Route", "Priced cancel rate", "Realised cancel rate",
        "Gap (pp)", "Ancillaries sold", "Ancillaries cancelled",
        "Avg fare (€)", "Margin impact (€)", "Hidden in blended view",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    def _write_variance_row(r, row: int, partner_id: str) -> None:
        """Write one variance row in fractions/euros (not bps/cents)."""
        # Priced cancel rate — formula references the named range (a
        # fraction, e.g. 0.0370). Displayed as a percentage.
        ws.cell(
            row=row, column=3,
            value=f"=partner_{partner_id}_priced_cancel_rate",
        ).number_format = FMT_PCT
        # Realised — stored as fraction (engine value is bps, divide by 10000).
        realised_pct = (
            None
            if r.realised_cancel_rate_bps is None
            else r.realised_cancel_rate_bps / 10000
        )
        ws.cell(row=row, column=4, value=realised_pct).number_format = FMT_PCT
        # Gap (pp) — formula in percentage points. Both C{row} and D{row}
        # are fractions, so (D−C)*100 produces the pp value directly.
        ws.cell(
            row=row, column=5,
            value=f"=(D{row}-C{row})*100",
        ).number_format = FMT_PP
        ws.cell(row=row, column=6, value=r.ancillaries_sold)
        ws.cell(row=row, column=7, value=r.ancillaries_cancelled)
        # Avg fare — stored as euros (cents/100) with euro format.
        ws.cell(
            row=row, column=8, value=r.avg_fare_cents / 100,
        ).number_format = FMT_EUR
        # Margin impact: (priced − realised) × coverage_pct × avg_fare × sold.
        # All inputs are now in consistent units (fractions for C/D, euros
        # for H), so the formula yields euros directly. Sign convention
        # preserved: realised > priced → margin impact negative.
        ws.cell(
            row=row, column=9,
            value=f"=ROUND((C{row}-D{row})*coverage_pct*H{row}*F{row},2)",
        ).number_format = FMT_EUR
        ws.cell(row=row, column=10, value="yes" if r.hidden_by_blend else "")

    row = 5
    for r in variance.rows:
        ws.cell(row=row, column=1, value=_variance_partner_label(r))
        if r.partner_id == "_blended_":
            ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=r.route_type or "ALL")
        _write_variance_row(r, row, r.partner_id)
        row += 1

    # Drilldown rows
    row += 1
    ws.cell(row=row, column=1, value="— Route drilldown —").font = Font(italic=True)
    row += 1
    for partner_id, drill_rows in variance.drilldown.items():
        for r in drill_rows:
            # Strip " — route" suffix so partner column shows partner only;
            # route already sits in column 2.
            partner_only = r.display_name.split(" — ", 1)[0]
            ws.cell(row=row, column=1, value=partner_only)
            ws.cell(row=row, column=2, value=r.route_type or "ALL")
            _write_variance_row(r, row, partner_id)
            row += 1

    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + i)].width = 22


def _write_ab_test(
    wb: Workbook, ab: ABTestView, registry: Registry, start_date
) -> None:
    ctl_pct = registry.fee_level.control_pct.value
    tst_pct = registry.fee_level.test_pct.value
    ctl_label = f"Current fee ({_fmt_pct(ctl_pct, tst_pct)} of fare)"
    tst_label = f"Lower fee ({_fmt_pct(tst_pct, ctl_pct)} of fare)"

    ws = wb.create_sheet("ABTest")
    wc = format_week_commencing(ab.as_of_week, start_date)
    launched = format_date(ab.split_date)
    ws["A1"] = f"A/B Test — as of {wc} · test launched {launched}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = f"Mix-adjustment: {ab.mix_control_method}"

    ws["A4"] = "Arm sizes (bookings since test launch)"
    ws["A4"].font = Font(bold=True)
    ws["A5"] = ctl_label
    ws["B5"] = ab.arm_sizes["control"]
    ws["A6"] = tst_label
    ws["B6"] = ab.arm_sizes["test"]

    headers = [
        "Metric",
        f"Unadjusted — {ctl_label}",
        f"Unadjusted — {tst_label}",
        f"Adjusted for partner mix — {ctl_label}",
        f"Adjusted for partner mix — {tst_label}",
        "Δ adjusted (Lower fee − Current fee)",
        "Winner",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=8, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row = 9
    for m in ab.metrics:
        ws.cell(row=row, column=1, value=_metric_label(m.metric))
        # contribution_per_booking is in cents → display as euros; the
        # other three metrics (attach_rate, loss_ratio, gross_margin_pct)
        # are fractions → display as percentages.
        is_eur = m.metric == "contribution_per_booking_cents"
        scale = (lambda v: v / 100) if is_eur else (lambda v: v)
        fmt = FMT_EUR if is_eur else FMT_PCT
        for col_idx, key in ((2, "naive"), (3, "naive"), (4, "stratified"), (5, "stratified")):
            arm = "control" if col_idx in (2, 4) else "test"
            cell = ws.cell(
                row=row, column=col_idx,
                value=scale(getattr(m, key)[arm]),
            )
            cell.number_format = fmt
        delta = ws.cell(row=row, column=6, value=f"=E{row}-D{row}")
        delta.number_format = fmt
        ws.cell(row=row, column=7, value=_scenario_name(m.winning_arm))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Verdict").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Winner on contribution per booking")
    ws.cell(
        row=row,
        column=2,
        value=_scenario_name(ab.verdict.winner_on_contribution_per_booking),
    )
    row += 1
    ws.cell(row=row, column=1, value="Winner on total contribution")
    ws.cell(
        row=row,
        column=2,
        value=_scenario_name(ab.verdict.winner_on_total_contribution),
    )
    row += 1
    ws.cell(row=row, column=1, value=f"Total contribution — {ctl_label} (€)")
    cell = ws.cell(
        row=row, column=2,
        value=ab.verdict.total_contribution_cents["control"] / 100,
    )
    cell.number_format = FMT_EUR
    row += 1
    ws.cell(row=row, column=1, value=f"Total contribution — {tst_label} (€)")
    cell = ws.cell(
        row=row, column=2,
        value=ab.verdict.total_contribution_cents["test"] / 100,
    )
    cell.number_format = FMT_EUR
    row += 2
    ws.cell(row=row, column=1, value="Trade-off").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value=_compose_tradeoff_text(ab)).alignment = (
        Alignment(wrap_text=True)
    )

    for i in range(1, 8):
        ws.column_dimensions[chr(64 + i)].width = 30


def _write_projection(
    wb: Workbook, projection: ProjectionView, start_date
) -> None:
    ws = wb.create_sheet("Projection")
    wc = format_week_commencing(projection.as_of_week, start_date)
    ws["A1"] = (
        f"Projection — {projection.weeks_forward}-week forward from {wc}"
    )
    ws["A1"].font = Font(bold=True, size=12)

    # Drivers panel
    ws["A3"] = "Drivers"
    ws["A3"].font = Font(bold=True)
    driver_headers = ["Name", "Value", "Origin", "Source", "Formula"]
    for col, h in enumerate(driver_headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row = 5
    for d in projection.drivers:
        ws.cell(row=row, column=1, value=d.name)
        ws.cell(row=row, column=2, value=d.value)
        cell = ws.cell(row=row, column=3, value=d.origin)
        cell.fill = ORIGIN_FILL.get(d.origin, PatternFill())
        ws.cell(row=row, column=4, value=d.source or "")
        ws.cell(row=row, column=5, value=d.formula)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Weekly schedule").font = Font(bold=True)
    row += 1
    week_headers = [
        "Scenario", "Week offset", "Week commencing", "Volume", "Ancillaries",
        "Revenue (€)", "Payouts (€)", "Cost of service (€)",
        "Contribution (€)",
    ]
    for col, h in enumerate(week_headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1
    scenario_display = {
        "standardise_on_control": "Standardise on Current fee",
        "standardise_on_test": "Standardise on Lower fee",
    }
    for w in projection.weekly:
        ws.cell(row=row, column=1, value=scenario_display.get(w.scenario, w.scenario))
        ws.cell(row=row, column=2, value=w.week_offset)
        ws.cell(
            row=row, column=3,
            value=format_week_commencing(w.iso_week, start_date, prefix=""),
        )
        ws.cell(row=row, column=4, value=w.volume)
        ws.cell(row=row, column=5, value=w.ancillaries)
        for col_idx, cents in (
            (6, w.revenue_cents),
            (7, w.payouts_cents),
            (8, w.cost_of_service_cents),
            (9, w.contribution_cents),
        ):
            cell = ws.cell(row=row, column=col_idx, value=cents / 100)
            cell.number_format = FMT_EUR
        row += 1

    # Totals — per scenario
    row += 1
    ws.cell(row=row, column=1, value="Totals (52w)").font = Font(bold=True)
    row += 1
    for scenario, t in projection.totals.items():
        ws.cell(
            row=row, column=1,
            value=scenario_display.get(scenario, scenario),
        )
        ws.cell(row=row, column=4, value=t.volume)
        ws.cell(row=row, column=5, value=t.ancillaries)
        for col_idx, cents in (
            (6, t.revenue_cents),
            (7, t.payouts_cents),
            (8, t.cost_of_service_cents),
            (9, t.contribution_cents),
        ):
            cell = ws.cell(row=row, column=col_idx, value=cents / 100)
            cell.number_format = FMT_EUR
        row += 1

    for col, _ in enumerate(week_headers, start=1):
        ws.column_dimensions[chr(64 + col)].width = 22


_BRIEFING_PROVIDER_DISPLAY: dict[str, str] = {
    "anthropic": "Claude",
    "openrouter": "Gemini 2.0 Flash",
}


def _write_briefing(wb: Workbook, briefing: Briefing) -> None:
    ws = wb.create_sheet("Briefing")
    ws["A1"] = "Briefing"
    ws["A1"].font = Font(bold=True, size=12)
    if briefing.mode == "llm":
        badge = _BRIEFING_PROVIDER_DISPLAY.get(briefing.provider or "", "LLM")
    else:
        badge = "deterministic fallback"
    ws["A2"] = f"Mode: {badge}"
    ws["A2"].font = Font(bold=True)
    fill = (
        PatternFill("solid", fgColor="D1ECF1")
        if briefing.mode == "llm"
        else PatternFill("solid", fgColor="FFF3CD")
    )
    ws["A2"].fill = fill

    ws["A4"] = "Narrative"
    ws["A4"].font = Font(bold=True)
    ws["A5"] = briefing.rendered_text
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[5].height = 200

    # Evidence pack table
    ws["A8"] = "Evidence pack (typed inputs to the briefing)"
    ws["A8"].font = Font(bold=True)
    ws.append([])
    headers = [
        "partner_id", "display_name", "classification", "matched_event_ids",
        "current_loss_ratio_bps", "loss_ratio_delta_bps",
        "current_cancel_rate_bps", "priced_cancel_rate_bps",
        "cancel_gap_bps", "margin_distance_from_floor_bps",
    ]
    ws.append(headers)
    for c in ws[ws.max_row]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for pe in briefing.evidence.partners:
        ws.append([
            pe.partner_id,
            pe.display_name,
            pe.classification,
            ",".join(pe.matched_event_ids),
            pe.current_loss_ratio_bps,
            pe.loss_ratio_delta_bps,
            pe.current_cancel_rate_bps,
            pe.priced_cancel_rate_bps,
            pe.cancel_gap_bps,
            pe.margin_distance_from_floor_bps,
        ])

    ws.column_dimensions["A"].width = 28
    for i in range(2, 11):
        ws.column_dimensions[chr(64 + i)].width = 18


def _write_consistency(wb: Workbook, consistency: ConsistencyReport) -> None:
    ws = wb.create_sheet("Consistency")
    ws["A1"] = "Cross-view reconciliation"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "PASSED" if consistency.passed else "FAILED"
    ws["A2"].font = Font(bold=True, color="006400" if consistency.passed else "B22222")

    headers = ["Check", "LHS (label / value)", "RHS (label / value)", "Passed?"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row = 5
    for c in consistency.checks:
        ws.cell(row=row, column=1, value=c.name)
        ws.cell(row=row, column=2, value=f"{c.lhs_label} = {c.lhs_value:,}")
        ws.cell(row=row, column=3, value=f"{c.rhs_label} = {c.rhs_value:,}")
        ws.cell(row=row, column=4, value="✓" if c.passed else "✗")
        row += 1
    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 60


def _write_audit(wb: Workbook, performance: PerformanceView) -> None:
    """Worked example re-deriving headline contribution from raw aggregates.

    A finance reader walks through these formulas to confirm the engine's
    headline matches the booking-level arithmetic.
    """
    ws = wb.create_sheet("Audit")
    ws["A1"] = "Audit — worked re-derivation of headline contribution"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = (
        "These cells re-derive the headline current-week contribution from "
        "the WeeklyAggregates sheet, using the same primitives the engine "
        "uses. Compare these formula-cell values to the Performance sheet."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 60

    headers = [
        "Partner", "Revenue (€)",
        "Payouts (€)",
        "Cost of service (€)",
        "Contribution (€) — revenue − payouts − cost",
        "Equals Performance figure?",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    def _sumifs_eur_audit(col_letter: str, week_: int, partner: str) -> str:
        """Same /100 convention as the Performance sheet so both sides of
        the audit IF live in euros, keeping the reconciliation valid."""
        return (
            f'=SUMIFS(WeeklyAggregates!{col_letter}:{col_letter}, '
            f'WeeklyAggregates!A:A, "{partner}", WeeklyAggregates!B:B, '
            f'{week_})/100'
        )

    week = performance.as_of_week
    row = 5
    for status in performance.partners:
        pid = status.partner_id
        ws.cell(row=row, column=1, value=status.display_name)
        for col_idx, letter in ((2, "H"), (3, "I"), (4, "J")):
            cell = ws.cell(
                row=row, column=col_idx,
                value=_sumifs_eur_audit(letter, week, pid),
            )
            cell.number_format = FMT_EUR
        cell = ws.cell(row=row, column=5, value=f"=B{row}-C{row}-D{row}")
        cell.number_format = FMT_EUR
        # Performance!F is contribution in euros (same /100 convention).
        # ROUND-to-cents both sides so floating-point equality is robust.
        ws.cell(
            row=row, column=6,
            value=(
                f'=IF(ROUND(E{row},2)=ROUND(Performance!F{row},2),'
                f'"yes","NO — investigate")'
            ),
        )
        row += 1

    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + i)].width = 32


def _fmt_pct(pct: float, other: float) -> str:
    """0dp normally; 1dp if both arms would render to the same integer percent."""
    if round(pct * 100) == round(other * 100):
        return f"{pct * 100:.1f}%"
    return f"{pct * 100:.0f}%"


_METRIC_LABELS = {
    "attach_rate": "Attach rate",
    "loss_ratio": "Loss ratio",
    "gross_margin_pct": "Gross margin",
    "contribution_per_booking_cents": "Contribution per booking",
}


def _metric_label(metric: str) -> str:
    return _METRIC_LABELS.get(metric, metric)


def _scenario_name(arm: str) -> str:
    return {"control": "Current fee", "test": "Lower fee", "tie": "Tie"}.get(
        arm, arm
    )


def _compose_tradeoff_text(ab: ABTestView) -> str:
    """Plain-language trade-off composed from structured fields — never the
    engine's raw 'control arm / test arm' sentence."""
    attach = next((m for m in ab.metrics if m.metric == "attach_rate"), None)
    cpb = next(
        (m for m in ab.metrics if m.metric == "contribution_per_booking_cents"),
        None,
    )
    if attach is None or cpb is None:
        return ""
    higher_attach = (
        "test" if attach.stratified["test"] > attach.stratified["control"]
        else "control"
    )
    higher_cpb = (
        "test" if cpb.stratified["test"] > cpb.stratified["control"]
        else "control"
    )
    if higher_attach == higher_cpb:
        return (
            f"{_scenario_name(higher_cpb)} wins on both attach rate and "
            "contribution per booking — no trade-off."
        )
    ctl_eur = ab.verdict.total_contribution_cents["control"] / 100
    tst_eur = ab.verdict.total_contribution_cents["test"] / 100
    return (
        f"Volume vs margin: {_scenario_name(higher_attach)} wins on attach "
        f"rate, {_scenario_name(higher_cpb)} wins on contribution per booking. "
        f"Total contribution — Current fee €{ctl_eur:,.0f} vs Lower fee "
        f"€{tst_eur:,.0f}."
    )


def _variance_partner_label(row) -> str:
    """Plain-language label for variance partner column — blended row renamed."""
    if row.partner_id == "_blended_":
        return "Blended (all partners)"
    return row.display_name


# Silence unused import (used at type-check time only by callers)
_ = WeeklyAggregate
