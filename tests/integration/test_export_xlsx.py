"""XLSX export verification (T059, FR-025)."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.config.loader import load_registry
from src.data.generator import generate_dataset
from src.engine.ab_test import compute_ab
from src.engine.briefing import compute_briefing
from src.engine.consistency import check_consistency
from src.engine.performance import compute_performance
from src.engine.projection import compute_projection
from src.engine.variance import compute_variance
from src.export.xlsx import write_workbook

REGISTRY_PATH = Path(__file__).resolve().parents[2] / "config" / "registry.yaml"


@pytest.fixture(scope="module")
def workbook_path(tmp_path_factory):
    registry = load_registry(REGISTRY_PATH)
    df = generate_dataset(registry, write_parquet=False)
    pv = compute_performance(registry, df)
    vv = compute_variance(registry, df)
    ab = compute_ab(registry, df)
    pj = compute_projection(registry, df, ab)
    briefing = compute_briefing(pv, registry, force_template=True)
    consistency = check_consistency(
        performance=pv, variance=vv, ab_test=ab, bookings=df,
        registry=registry, projection=pj,
    )
    out = tmp_path_factory.mktemp("exports") / "DA_Engine_test.xlsx"
    write_workbook(
        registry=registry,
        bookings_df=df,
        performance=pv,
        variance=vv,
        ab_test=ab,
        projection=pj,
        briefing=briefing,
        consistency=consistency,
        path=out,
    )
    return out


def test_workbook_loads_and_has_expected_sheets(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path, data_only=False)
    expected = {
        "README", "Assumptions", "WeeklyAggregates",
        "Performance", "Variance", "ABTest", "Projection",
        "Briefing", "Consistency", "Audit",
    }
    assert expected.issubset(set(wb.sheetnames))


def test_named_ranges_defined_for_every_registry_leaf(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path, data_only=False)
    names = set(wb.defined_names)
    must_have = {
        "coverage_pct",
        "payment_processing_pct",
        "servicing_cost_per_unit_cents",
        "fee_level_control_pct",
        "fee_level_test_pct",
        "margin_floor_bps",
        "classification_material_gap_bps",
        "projection_trend_factor",
    }
    assert must_have.issubset(names), f"missing names: {must_have - names}"
    # Per-partner named ranges
    assert any(n.startswith("partner_") and n.endswith("_priced_cancel_rate") for n in names)
    # Feature 002 — legacy named ranges removed
    assert "fee_level_control_cents" not in names
    assert "fee_level_test_cents" not in names


def test_variance_sheet_uses_named_ranges_in_formulas(workbook_path: Path) -> None:
    """Derived cells must contain live formulas, not pre-evaluated constants
    (FR-025, SC-004)."""
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb["Variance"]
    formula_cells = 0
    name_refs = 0
    for row in ws.iter_rows(min_row=5):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells += 1
                if "coverage_pct" in cell.value or "partner_" in cell.value:
                    name_refs += 1
    assert formula_cells > 0, "Variance has no formula cells"
    assert name_refs > 0, "No Variance formula references a named range"


def test_audit_sheet_has_check_formulas(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb["Audit"]
    formula_cells = sum(
        1
        for row in ws.iter_rows(min_row=5)
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    # Per-partner: 5 formulas (revenue, payouts, cos, contribution, check)
    # × 3 partners = 15
    assert formula_cells >= 12, f"Audit only has {formula_cells} formulas"


def test_briefing_sheet_carries_mode_badge(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb["Briefing"]
    badge_text = str(ws["A2"].value or "")
    assert "Mode:" in badge_text
    assert (
        "deterministic fallback" in badge_text
        or "Claude" in badge_text
        or "Gemini" in badge_text
    )


def test_workbook_forces_full_calc_on_load(workbook_path: Path) -> None:
    """openpyxl writes formulas without cached values; some viewers leave
    cells blank until a manual recalc. fullCalcOnLoad makes Excel/Sheets/
    preview compute the workbook on open."""
    wb = load_workbook(workbook_path, data_only=False)
    assert wb.calculation.fullCalcOnLoad is True


def test_performance_currency_columns_use_euro_format(workbook_path: Path) -> None:
    """Revenue/Payouts/CoS/Contribution must display as euros, not raw cents."""
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb["Performance"]
    # Header text changed to "(€)"
    headers = [ws.cell(row=4, column=c).value for c in range(1, 11)]
    assert "Revenue (€)" in headers
    assert "Contribution (€)" in headers
    assert not any("(cents)" in str(h or "") for h in headers)
    # Number format on the first partner row's currency cells
    for col_idx in (3, 4, 5, 6):
        fmt = ws.cell(row=5, column=col_idx).number_format
        assert "€" in fmt, (
            f"Performance!{ws.cell(row=5, column=col_idx).coordinate} "
            f"missing euro number format (got {fmt!r})"
        )


def test_performance_status_and_margin_distance_are_formulas(
    workbook_path: Path,
) -> None:
    """Status + margin distance must be live formulas referencing the
    Assumptions named ranges, not hardcoded literals."""
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb["Performance"]
    status_cell = ws.cell(row=5, column=2).value
    distance_cell = ws.cell(row=5, column=10).value
    assert isinstance(status_cell, str) and status_cell.startswith("=")
    assert "margin_approaching_floor_buffer_bps" in status_cell
    assert isinstance(distance_cell, str) and distance_cell.startswith("=")
    assert "margin_floor_bps" in distance_cell


def test_variance_uses_percent_and_pp_not_bps(workbook_path: Path) -> None:
    """Variance sheet must match the app: % for rates, pp for gaps,
    euros for monetary."""
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb["Variance"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, 11)]
    assert "Priced cancel rate" in headers
    assert "Realised cancel rate" in headers
    assert "Gap (pp)" in headers
    assert "Avg fare (€)" in headers
    assert "Margin impact (€)" in headers
    # No stray "(bps)" / "(cents)" survivors
    assert not any("(bps)" in str(h or "") for h in headers)
    assert not any("(cents)" in str(h or "") for h in headers)
    # Spot-check formats on row 5
    assert ws.cell(row=5, column=3).number_format == "0.00%"
    assert ws.cell(row=5, column=4).number_format == "0.00%"
    assert "pp" in ws.cell(row=5, column=5).number_format
    assert "€" in ws.cell(row=5, column=8).number_format
    assert "€" in ws.cell(row=5, column=9).number_format


def test_consistency_sheet_lists_checks(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb["Consistency"]
    # PASSED/FAILED status at A2
    assert ws["A2"].value in ("PASSED", "FAILED")
    # Multiple check rows below header at row 4
    check_count = sum(
        1 for row in ws.iter_rows(min_row=5) if row[0].value is not None
    )
    assert check_count >= 8  # at least perf↔variance checks
